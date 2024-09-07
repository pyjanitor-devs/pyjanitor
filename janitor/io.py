from __future__ import annotations

import inspect
import os
import subprocess
import warnings
from collections import defaultdict
from glob import glob
from io import StringIO
from itertools import chain
from typing import IO, TYPE_CHECKING, Any, Iterable, Mapping, Union

import pandas as pd

from janitor.utils import (
    check,
    deprecated_alias,
    find_stack_level,
    import_message,
)

from .errors import JanitorError

warnings.simplefilter("always", DeprecationWarning)


@deprecated_alias(seperate_df="separate_df", filespath="files_path")
def read_csvs(
    files_path: Union[str, Iterable[str]],
    separate_df: bool = False,
    **kwargs: Any,
) -> Union[pd.DataFrame, dict]:
    """Read multiple CSV files and return a dictionary of DataFrames, or
    one concatenated DataFrame.

    Args:
        files_path: The filepath pattern matching the CSV files.
            Accepts regular expressions, with or without `.csv` extension.
            Also accepts iterable of file paths.
        separate_df: If `False` (default), returns a single Dataframe
            with the concatenation of the csv files.
            If `True`, returns a dictionary of separate DataFrames
            for each CSV file.
        **kwargs: Keyword arguments to pass into the
            original pandas `read_csv`.

    Raises:
        JanitorError: If `None` provided for `files_path`.
        JanitorError: If length of `files_path` is `0`.
        ValueError: If no CSV files exist in `files_path`.
        ValueError: If columns in input CSV files do not match.

    Returns:
        DataFrame of concatenated DataFrames or dictionary of DataFrames.
    """
    # Sanitize input
    if files_path is None:
        raise JanitorError("`None` provided for `files_path`")
    if not files_path:
        raise JanitorError("0 length `files_path` provided")

    # Read the csv files
    # String to file/folder or file pattern provided
    if isinstance(files_path, str):
        dfs_dict = {
            os.path.basename(f): pd.read_csv(f, **kwargs)
            for f in glob(files_path)
        }
    # Iterable of file paths provided
    else:
        dfs_dict = {
            os.path.basename(f): pd.read_csv(f, **kwargs) for f in files_path
        }
    # Check if dataframes have been read
    if not dfs_dict:
        raise ValueError("No CSV files to read with the given `files_path`")
    # Concatenate the dataframes if requested (default)
    col_names = list(dfs_dict.values())[0].columns  # noqa: PD011
    if not separate_df:
        # If columns do not match raise an error
        for df in dfs_dict.values():  # noqa: PD011
            if not all(df.columns == col_names):
                raise ValueError(
                    "Columns in input CSV files do not match."
                    "Files cannot be concatenated."
                )
        return pd.concat(
            list(dfs_dict.values()),
            ignore_index=True,
            sort=False,  # noqa: PD011
            copy=False,
        )
    return dfs_dict


def read_commandline(
    cmd: str, engine: str = "pandas", **kwargs: Any
) -> Mapping:
    """Read a CSV file based on a command-line command.

    For example, you may wish to run the following command on `sep-quarter.csv`
    before reading it into a pandas DataFrame:

    ```bash
    cat sep-quarter.csv | grep .SEA1AA
    ```

    In this case, you can use the following Python code to load the dataframe:

    ```python
    import janitor as jn
    df = jn.read_commandline("cat data.csv | grep .SEA1AA")
    ```

    This function assumes that your command line command will return
    an output that is parsable using the relevant engine and StringIO.
    This function defaults to using `pd.read_csv` underneath the hood.
    Keyword arguments are passed through as-is.

    Args:
        cmd: Shell command to preprocess a file on disk.
        engine: DataFrame engine to process the output of the shell command.
            Currently supports both pandas and polars.
        **kwargs: Keyword arguments that are passed through to
            the engine's csv reader.


    Returns:
        A DataFrame parsed from the stdout of the underlying
            shell.
    """

    check("cmd", cmd, [str])
    if engine not in {"pandas", "polars"}:
        raise ValueError("engine should be either pandas or polars.")
    # adding check=True ensures that an explicit, clear error
    # is raised, so that the user can see the reason for the failure
    outcome = subprocess.run(
        cmd, shell=True, capture_output=True, text=True, check=True
    )
    if engine == "polars":
        try:
            import polars as pl
        except ImportError:
            import_message(
                submodule="polars",
                package="polars",
                conda_channel="conda-forge",
                pip_install=True,
            )
        return pl.read_csv(StringIO(outcome.stdout), **kwargs)
    return pd.read_csv(StringIO(outcome.stdout), **kwargs)


if TYPE_CHECKING:
    from openpyxl import Workbook


def xlsx_table(
    path: Union[str, IO, Workbook],
    sheetname: str = None,
    table: Union[str, list, tuple] = None,
    engine: str = "pandas",
) -> Mapping:
    """Returns a DataFrame of values in a table in the Excel file.

    This applies to an Excel file, where the data range is explicitly
    specified as a Microsoft Excel table.

    If there is a single table in the sheet, or a string is provided
    as an argument to the `table` parameter, a DataFrame is returned;
    if there is more than one table in the sheet,
    and the `table` argument is `None`, or a list/tuple of names,
    a dictionary of DataFrames is returned, where the keys of the dictionary
    are the table names.

    Examples:
        >>> import pandas as pd
        >>> import polars as pl
        >>> from janitor import xlsx_table
        >>> filename="../pyjanitor/tests/test_data/016-MSPTDA-Excel.xlsx"

        Single table:

        >>> xlsx_table(filename, table='dCategory')
           CategoryID       Category
        0           1       Beginner
        1           2       Advanced
        2           3      Freestyle
        3           4    Competition
        4           5  Long Distance

        >>> xlsx_table(filename, table='dCategory', engine='polars')
        shape: (5, 2)
        ┌────────────┬───────────────┐
        │ CategoryID ┆ Category      │
        │ ---        ┆ ---           │
        │ i64        ┆ str           │
        ╞════════════╪═══════════════╡
        │ 1          ┆ Beginner      │
        │ 2          ┆ Advanced      │
        │ 3          ┆ Freestyle     │
        │ 4          ┆ Competition   │
        │ 5          ┆ Long Distance │
        └────────────┴───────────────┘

        Multiple tables:

        >>> out=xlsx_table(filename, table=["dCategory", "dSalesReps"])
        >>> out["dCategory"]
           CategoryID       Category
        0           1       Beginner
        1           2       Advanced
        2           3      Freestyle
        3           4    Competition
        4           5  Long Distance
        >>> out["dSalesReps"].head(3)
           SalesRepID             SalesRep Region
        0           1  Sioux Radcoolinator     NW
        1           2        Tyrone Smithe     NE
        2           3         Chantel Zoya     SW

    Args:
        path: Path to the Excel File. It can also be an openpyxl Workbook.
        table: Name of a table, or list of tables in the sheet.
        engine: DataFrame engine. Should be either pandas or polars.
            Defaults to pandas

    Raises:
        AttributeError: If a workbook is provided, and is a ReadOnlyWorksheet.
        ValueError: If there are no tables in the sheet.
        KeyError: If the provided table does not exist in the sheet.

    Returns:
        A DataFrame, or a dictionary of DataFrames,
            if there are multiple arguments for the `table` parameter,
            or the argument to `table` is `None`.
    """  # noqa : E501

    try:
        from openpyxl import load_workbook
        from openpyxl.workbook.workbook import Workbook
    except ImportError:
        import_message(
            submodule="io",
            package="openpyxl",
            conda_channel="conda-forge",
            pip_install=True,
        )
    # TODO: remove in version 1.0
    if sheetname:
        warnings.warn(
            "The keyword argument "
            "'sheetname' of 'xlsx_tables' is deprecated.",
            DeprecationWarning,
            stacklevel=find_stack_level(),
        )
    if engine not in {"pandas", "polars"}:
        raise ValueError("engine should be one of pandas or polars.")
    base_engine = pd
    if engine == "polars":
        try:
            import polars as pl

            base_engine = pl
        except ImportError:
            import_message(
                submodule="polars",
                package="polars",
                conda_channel="conda-forge",
                pip_install=True,
            )

    if table is not None:
        check("table", table, [str, list, tuple])
        if isinstance(table, (list, tuple)):
            for num, entry in enumerate(table):
                check(f"entry{num} in the table argument", entry, [str])
    if isinstance(path, Workbook):
        ws = path
    else:
        ws = load_workbook(
            filename=path, read_only=False, keep_links=False, data_only=True
        )
    if ws.read_only:
        raise ValueError("xlsx_table does not work in read only mode.")

    def _create_dataframe_or_dictionary_from_table(
        table_name_and_worksheet: tuple,
    ):
        """
        Create DataFrame/dictionary if table exists in Workbook
        """
        dictionary = {}
        for table_name, worksheet in table_name_and_worksheet:
            contents = worksheet.tables[table_name]
            header_exist = contents.headerRowCount
            coordinates = contents.ref
            data = worksheet[coordinates]
            if header_exist:
                header, *data = data
                header = [cell.value for cell in header]
            else:
                header = [f"C{num}" for num in range(len(data[0]))]
            data = zip(*data)
            data = ([entry.value for entry in cell] for cell in data)
            data = dict(zip(header, data))
            dictionary[table_name] = base_engine.DataFrame(data)
        return dictionary

    worksheets = [worksheet for worksheet in ws if worksheet.tables.items()]
    if not any(worksheets):
        raise ValueError("There are no tables in the Workbook.")
    table_is_a_string = False
    if table:
        if isinstance(table, str):
            table_is_a_string = True
            table = [table]
        table_names = (
            entry for worksheet in worksheets for entry in worksheet.tables
        )
        missing = set(table).difference(table_names)
        if missing:
            raise KeyError(f"Tables {*missing,} do not exist in the Workbook.")
        tables = [
            (entry, worksheet)
            for worksheet in worksheets
            for entry in worksheet.tables
            if entry in table
        ]
    else:
        tables = [
            (entry, worksheet)
            for worksheet in worksheets
            for entry in worksheet.tables
        ]
    data = _create_dataframe_or_dictionary_from_table(
        table_name_and_worksheet=tables
    )
    if table_is_a_string:
        return data[table[0]]
    return data


def xlsx_cells(
    path: Union[str, Workbook],
    sheetnames: Union[str, list, tuple] = None,
    start_point: Union[str, int] = None,
    end_point: Union[str, int] = None,
    read_only: bool = True,
    include_blank_cells: bool = True,
    fill: bool = False,
    font: bool = False,
    alignment: bool = False,
    border: bool = False,
    protection: bool = False,
    comment: bool = False,
    engine: str = "pandas",
    **kwargs: Any,
) -> Mapping:
    """Imports data from spreadsheet without coercing it into a rectangle.

    Each cell is represented by a row in a dataframe, and includes the
    cell's coordinates, the value, row and column position.
    The cell formatting (fill, font, border, etc) can also be accessed;
    usually this is returned as a dictionary in the cell, and the specific
    cell format attribute can be accessed using `pd.Series.str.get`
    or `pl.struct.field` if it is a polars DataFrame.

    Inspiration for this comes from R's [tidyxl][link] package.
    [link]: https://nacnudus.github.io/tidyxl/reference/tidyxl.html

    Examples:
        >>> import pandas as pd
        >>> import polars as pl
        >>> from janitor import xlsx_cells
        >>> pd.set_option("display.max_columns", None)
        >>> pd.set_option("display.expand_frame_repr", False)
        >>> pd.set_option("max_colwidth", None)
        >>> filename = "../pyjanitor/tests/test_data/worked-examples.xlsx"

        Each cell is returned as a row:

        >>> xlsx_cells(filename, sheetnames="highlights")
            value internal_value coordinate  row  column data_type  is_date number_format
        0     Age            Age         A1    1       1         s    False       General
        1  Height         Height         B1    1       2         s    False       General
        2       1              1         A2    2       1         n    False       General
        3       2              2         B2    2       2         n    False       General
        4       3              3         A3    3       1         n    False       General
        5       4              4         B3    3       2         n    False       General
        6       5              5         A4    4       1         n    False       General
        7       6              6         B4    4       2         n    False       General

        Access cell formatting such as fill:

        >>> out=xlsx_cells(filename, sheetnames="highlights", fill=True).select("value", "fill", axis='columns')
        >>> out
            value                                                                                                                                              fill
        0     Age     {'patternType': None, 'fgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}}
        1  Height     {'patternType': None, 'fgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}}
        2       1     {'patternType': None, 'fgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}}
        3       2     {'patternType': None, 'fgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}}
        4       3  {'patternType': 'solid', 'fgColor': {'rgb': 'FFFFFF00', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': 'FFFFFF00', 'type': 'rgb', 'tint': 0.0}}
        5       4  {'patternType': 'solid', 'fgColor': {'rgb': 'FFFFFF00', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': 'FFFFFF00', 'type': 'rgb', 'tint': 0.0}}
        6       5     {'patternType': None, 'fgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}}
        7       6     {'patternType': None, 'fgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}, 'bgColor': {'rgb': '00000000', 'type': 'rgb', 'tint': 0.0}}

        Specific cell attributes can be accessed by using Pandas' `series.str.get`:

        >>> out.fill.str.get("fgColor").str.get("rgb")
        0    00000000
        1    00000000
        2    00000000
        3    00000000
        4    FFFFFF00
        5    FFFFFF00
        6    00000000
        7    00000000
        Name: fill, dtype: object

        Access cell formatting in a polars DataFrame:

        >>> out = xlsx_cells(filename, sheetnames="highlights", engine='polars', fill=True).get_column('fill')
        >>> out
        shape: (8,)
        Series: 'fill' [struct[3]]
        [
           {null,{"00000000","rgb",0.0},{"00000000","rgb",0.0}}
           {null,{"00000000","rgb",0.0},{"00000000","rgb",0.0}}
           {null,{"00000000","rgb",0.0},{"00000000","rgb",0.0}}
           {null,{"00000000","rgb",0.0},{"00000000","rgb",0.0}}
           {"solid",{"FFFFFF00","rgb",0.0},{"FFFFFF00","rgb",0.0}}
           {"solid",{"FFFFFF00","rgb",0.0},{"FFFFFF00","rgb",0.0}}
           {null,{"00000000","rgb",0.0},{"00000000","rgb",0.0}}
           {null,{"00000000","rgb",0.0},{"00000000","rgb",0.0}}
        ]

        Specific cell attributes can be acessed via Polars' struct:

        >>> out.struct.field('fgColor').struct.field('rgb')
        shape: (8,)
        Series: 'rgb' [str]
        [
           "00000000"
           "00000000"
           "00000000"
           "00000000"
           "FFFFFF00"
           "FFFFFF00"
           "00000000"
           "00000000"
        ]


    Args:
        path: Path to the Excel File. It can also be an openpyxl Workbook.
        sheetnames: Names of the sheets from which the cells are to be extracted.
            If `None`, all the sheets in the file are extracted;
            if it is a string, or list or tuple, only the specified sheets are extracted.
        start_point: Start coordinates of the Excel sheet. This is useful
            if the user is only interested in a subsection of the sheet.
            If `start_point` is provided, `end_point` must be provided as well.
        end_point: End coordinates of the Excel sheet. This is useful
            if the user is only interested in a subsection of the sheet.
            If `end_point` is provided, `start_point` must be provided as well.
        read_only: Determines if the entire file is loaded in memory,
            or streamed. For memory efficiency, read_only should be set to `True`.
            Some cell properties like `comment`, can only be accessed by
            setting `read_only` to `False`.
        include_blank_cells: Determines if cells without a value should be included.
        fill: If `True`, return fill properties of the cell.
            It is usually returned as a dictionary.
        font: If `True`, return font properties of the cell.
            It is usually returned as a dictionary.
        alignment: If `True`, return alignment properties of the cell.
            It is usually returned as a dictionary.
        border: If `True`, return border properties of the cell.
            It is usually returned as a dictionary.
        protection: If `True`, return protection properties of the cell.
            It is usually returned as a dictionary.
        comment: If `True`, return comment properties of the cell.
            It is usually returned as a dictionary.
        engine: DataFrame engine. Should be either pandas or polars.
        **kwargs: Any other attributes of the cell, that can be accessed from openpyxl.

    Raises:
        ValueError: If kwargs is provided, and one of the keys is a default column.
        AttributeError: If kwargs is provided and any of the keys
            is not a openpyxl cell attribute.

    Returns:
        A DataFrame, or a dictionary of DataFrames.
    """  # noqa : E501

    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import Cell
        from openpyxl.cell.read_only import ReadOnlyCell
        from openpyxl.workbook.workbook import Workbook
    except ImportError:
        import_message(
            submodule="io",
            package="openpyxl",
            conda_channel="conda-forge",
            pip_install=True,
        )

    path_is_workbook = isinstance(path, Workbook)
    if not path_is_workbook:
        # for memory efficiency, read_only is set to True
        # if comments is True, read_only has to be False,
        # as lazy loading is not enabled for comments
        if comment and read_only:
            raise ValueError(
                "To access comments, kindly set 'read_only' to False."
            )
        path = load_workbook(
            filename=path, read_only=read_only, keep_links=False
        )
    if engine not in {"pandas", "polars"}:
        raise ValueError("engine should be one of pandas or polars.")
    base_engine = pd
    if engine == "polars":
        try:
            import polars as pl

            base_engine = pl
        except ImportError:
            import_message(
                submodule="polars",
                package="polars",
                conda_channel="conda-forge",
                pip_install=True,
            )
    # start_point and end_point applies if the user is interested in
    # only a subset of the Excel File and knows the coordinates
    if start_point or end_point:
        check("start_point", start_point, [str, int])
        check("end_point", end_point, [str, int])

    defaults = (
        "value",
        "internal_value",
        "coordinate",
        "row",
        "column",
        "data_type",
        "is_date",
        "number_format",
    )

    parameters = {
        "fill": fill,
        "font": font,
        "alignment": alignment,
        "border": border,
        "protection": protection,
        "comment": comment,
    }

    if kwargs:
        if path_is_workbook:
            if path.read_only:
                _cell = ReadOnlyCell
            else:
                _cell = Cell
        else:
            if read_only:
                _cell = ReadOnlyCell
            else:
                _cell = Cell

        attrs = {
            attr
            for attr, _ in inspect.getmembers(_cell, not (inspect.isroutine))
            if not attr.startswith("_")
        }

        for key in kwargs:
            if key in defaults:
                raise ValueError(
                    f"{key} is part of the default attributes "
                    "returned as a column."
                )
            elif key not in attrs:
                raise AttributeError(
                    f"{key} is not a recognized attribute of {_cell}."
                )
        parameters.update(kwargs)

    if not sheetnames:
        sheetnames = path.sheetnames
    elif isinstance(sheetnames, str):
        sheetnames = [sheetnames]
    else:
        check("sheetnames", sheetnames, [str, list, tuple])

    out = {
        sheetname: _xlsx_cells(
            path[sheetname],
            defaults,
            parameters,
            start_point,
            end_point,
            include_blank_cells,
            base_engine=base_engine,
        )
        for sheetname in sheetnames
    }
    if len(out) == 1:
        _, out = out.popitem()

    if (not path_is_workbook) and path.read_only:
        path.close()

    return out


def _xlsx_cells(
    wb: Workbook,
    defaults: tuple,
    parameters: dict,
    start_point: Union[str, int],
    end_point: Union[str, int],
    include_blank_cells: bool,
    base_engine,
):
    """
    Function to process a single sheet. Returns a DataFrame.

    Args:
        wb: Openpyxl Workbook.
        defaults: Sequence of default cell attributes.
        parameters: Dictionary of cell attributes to be retrieved.
            that will always be returned as columns.
        start_point: start coordinates of the Excel sheet.
        end_point: end coordinates of the Excel sheet.
        include_blank_cells: Determines if empty cells should be included.
        path_is_workbook: True/False.

    Returns:
        A DataFrame.
    """

    if start_point:
        wb = wb[start_point:end_point]
    wb = chain.from_iterable(wb)
    frame = defaultdict(list)

    for cell in wb:
        if (cell.value is None) and (not include_blank_cells):
            continue
        for value in defaults:
            outcome = getattr(cell, value, None)
            if value.startswith("is_"):
                pass
            elif outcome is not None:
                outcome = str(outcome)
            frame[value].append(outcome)
        for parent, boolean_value in parameters.items():
            check(f"The value for {parent}", boolean_value, [bool])
            if not boolean_value:
                continue
            boolean_value = _object_to_dict(getattr(cell, parent, None))
            if isinstance(boolean_value, dict) or (boolean_value is None):
                pass
            else:
                boolean_value = str(boolean_value)
            frame[parent].append(boolean_value)
    return base_engine.DataFrame(frame)


def _object_to_dict(obj):
    """
    Recursively get the attributes
    of an object as a dictionary.

    Args:
        obj: Object whose attributes are to be extracted.

    Returns:
        A dictionary or the object.
    """
    # https://stackoverflow.com/a/71366813
    data = {}
    if getattr(obj, "__dict__", None):
        for key, value in obj.__dict__.items():
            data[key] = _object_to_dict(value)
        return data
    return obj
