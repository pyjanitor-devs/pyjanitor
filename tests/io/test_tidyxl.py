import pytest

from janitor import io
from pathlib import Path
from openpyxl import load_workbook
from pandas.testing import assert_series_equal


TEST_DATA_DIR = "tests/test_data"
filename = Path(TEST_DATA_DIR).joinpath("worked-examples.xlsx").resolve()


wb = load_workbook(filename)
wb_c = load_workbook(filename, read_only=True)


def test_sheetname_type():
    """Raise error if sheet name is the wrong type."""
    with pytest.raises(TypeError, match="sheetnames should be one of .+"):
        io.xlsx_cells(filename, 5)


def test_sheetname_presence():
    """Raise error if sheet name does not exist."""
    with pytest.raises(KeyError, match="Worksheet 5 .+"):
        io.xlsx_cells(filename, [5])


def test_comment_read_only():
    """Raise error if comment and read_only is True."""
    with pytest.raises(ValueError, match="To access comments.+"):
        io.xlsx_cells(filename, "clean", comment=True, read_only=True)


def test_check_filename():
    """Raise error if file does not exist."""
    with pytest.raises(
        FileNotFoundError, match=r"\[Errno 2\] No such file or directory.+"
    ):
        io.xlsx_cells("excel.xlsx", "clean")


def test_check_start_none():
    """
    Raise error if start_point is None,
    and end_point is not None.
    """
    with pytest.raises(TypeError, match="start_point should be one.+"):
        io.xlsx_cells(filename, "clean", start_point=None, end_point="B5")


def test_check_end_none():
    """
    Raise error if start_point is not None,
    and end_point is None.
    """
    with pytest.raises(TypeError, match="end_point should be one.+"):
        io.xlsx_cells(wb, "clean", start_point="A1", end_point=None)


def test_output_kwargs_defaults():
    """
    Raise AttributeError if kwargs is provided
    and the key is already part of the default attributes
    that are returned as columns.
    """
    with pytest.raises(
        ValueError, match="internal_value is part of the default attributes.+"
    ):
        io.xlsx_cells(wb, "clean", internal_value=True)


def test_output_kwargs_defaults_read_only_true():
    """
    Raise AttributeError if kwargs is provided
    and the key is already part of the default attributes
    that are returned as columns.
    """
    with pytest.raises(
        ValueError, match="internal_value is part of the default attributes.+"
    ):
        io.xlsx_cells(wb_c, "clean", internal_value=True)


def test_output_wrong_kwargs():
    """
    Raise AttributeError if kwargs is provided
    and the key is not a valid openpyxl cell attribute.
    """
    with pytest.raises(
        AttributeError, match="font_color is not a recognized attribute of.+"
    ):
        io.xlsx_cells(filename, "clean", font_color=True)


def test_output_wrong_kwargs_read_only_false():
    """
    Raise AttributeError if kwargs is provided
    and the key is not a valid openpyxl cell attribute.
    """
    with pytest.raises(
        AttributeError, match="font_color is not a recognized attribute of.+"
    ):
        io.xlsx_cells(filename, "clean", font_color=True, read_only=False)


def test_default_values():
    """
    Test output for default values.
    """
    assert "Matilda" in io.xlsx_cells(wb, "clean")["value"].tolist()


def test_default_values_blank_cells_true():
    """
    Test output for default values if include_blank_cells.
    """
    assert None in io.xlsx_cells(filename, "pivot-notes")["value"].tolist()


def test_default_values_blank_cells_false():
    """
    Test output for default values if include_blank_cells is False.
    """
    assert (
        None
        not in io.xlsx_cells(
            wb,
            "pivot-notes",
            include_blank_cells=False,
            start_point="A1",
            end_point="G7",
        )["value"].tolist()
    )


def test_output_parameters():
    """Test output for existing parameters."""
    result = (
        io.xlsx_cells(filename, "pivot-notes", font=True)["font"]
        .str.get("name")
        .tolist()
    )
    assert "Calibri" in result


def test_output_kwargs():
    """Test output for extra attributes via kwargs."""
    result = io.xlsx_cells(wb, "pivot-notes", col_idx=True)
    assert_series_equal(result["column"], result["col_idx"].rename("column"))


def test_output_kwargs_type():
    """Test output for parameters if value is not boolean."""
    with pytest.raises(TypeError):
        io.xlsx_cells(wb, "pivot-notes", col_idx="True")


def test_output_sheetnames_sequence():
    """Test output if sheetnames is a list."""
    result = (
        io.xlsx_cells(filename, sheetnames=["pivot-notes"], font=True)["font"]
        .str.get("name")
        .tolist()
    )
    assert "Calibri" in result


def test_output_sheetnames_none():
    """Test output if sheetnames is None."""
    result = (
        io.xlsx_cells(wb, sheetnames=None, font=True)["pivot-notes"]["font"]
        .str.get("name")
        .tolist()
    )
    assert "Calibri" in result


wb_c.close()
