import re
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from pandas.testing import assert_frame_equal

from janitor import io

TEST_DATA_DIR = "tests/test_data"
filename = Path(TEST_DATA_DIR).joinpath("016-MSPTDA-Excel.xlsx").resolve()
no_headers = (
    Path(
        TEST_DATA_DIR,
    )
    .joinpath("excel_without_headers.xlsx")
    .resolve()
)

no_tables = (
    Path(
        TEST_DATA_DIR,
    )
    .joinpath("file_example_XLSX_10.xlsx")
    .resolve()
)


@pytest.mark.xfail(reason="sheetname parameter deprecated.")
def test_check_sheetname():
    """Raise KeyError if sheetname does not exist."""
    with pytest.raises(KeyError):
        io.xlsx_table(filename, table=None)


def test_check_sheetname_warning():
    """Raise Warning if sheetnameis provided."""
    with pytest.warns(DeprecationWarning):
        io.xlsx_table(filename, sheetname="rar")


def test_check_filename():
    """Raise error if file does not exist."""
    with pytest.raises(FileNotFoundError):
        io.xlsx_table("excel.xlsx", table=None)


def test_table_exists():
    """Raise error if there is no table in the workbook."""
    with pytest.raises(
        ValueError, match="There are no tables in the Workbook."
    ):
        io.xlsx_table(no_tables, table="Cover")


def test_check_table_name_str():
    """Raise error if table name is not a string."""
    with pytest.raises(TypeError, match="table should be one of.+"):
        io.xlsx_table(filename, table=1)


def test_check_table_name_list():
    """Raise error if table name is not a string."""
    with pytest.raises(
        TypeError, match="entry0 in the table argument should be one of.+"
    ):
        io.xlsx_table(filename, table=[1, "rar"])


def test_table_name():
    """
    Raise error if `table` is not None,
    and the table name cannot be found.
    """
    with pytest.raises(
        KeyError,
        match=re.escape("Tables ('fake',) do not exist in the Workbook."),
    ):
        io.xlsx_table(filename, table="fake")


def test_wb_read_only():
    """
    Raise error if Workbook is provided, and read_only is True.
    """
    wb = load_workbook(filename, read_only=True)
    with pytest.raises(
        ValueError,
        match="xlsx_table does not work in read only mode.",
    ):
        io.xlsx_table(wb)
    wb.close()


def test_table_str():
    """Test output for single table."""
    expected = io.xlsx_table(filename, table="dSupplier")
    actual = (
        pd.read_excel(
            filename, engine="openpyxl", sheet_name="Tables", usecols="N:R"
        )
        .rename(columns={"SupplierID.1": "SupplierID"})
        .dropna()
    )
    assert_frame_equal(expected, actual)


def test_table_no_header():
    """Test output for single table, without header."""
    expected = io.xlsx_table(no_headers, table="dSalesReps")
    actual = pd.read_excel(
        no_headers,
        engine="openpyxl",
        sheet_name="Tables",
        usecols="A:C",
        names=["C0", "C1", "C2"],
    )
    assert_frame_equal(expected, actual)


def test_tables():
    """Test output for multiple tables."""
    wb = load_workbook(filename, read_only=False)
    expected = io.xlsx_table(wb, table=("dSalesReps", "dSupplier"))
    actual = {
        "dSalesReps": pd.read_excel(
            filename, engine="openpyxl", sheet_name="Tables", usecols="A:C"
        ),
        "dSupplier": pd.read_excel(
            filename, engine="openpyxl", sheet_name="Tables", usecols="N:R"
        )
        .rename(columns={"SupplierID.1": "SupplierID"})
        .dropna(),
    }
    for key, value in expected.items():
        assert_frame_equal(value, actual[key])


def test_tables_none():
    """Test output for multiple tables."""
    expected = io.xlsx_table(filename)
    actual = {
        "dSalesReps": pd.read_excel(
            filename, engine="openpyxl", sheet_name="Tables", usecols="A:C"
        ),
        "dSupplier": pd.read_excel(
            filename, engine="openpyxl", sheet_name="Tables", usecols="N:R"
        )
        .rename(columns={"SupplierID.1": "SupplierID"})
        .dropna(),
        "dProduct": pd.read_excel(
            filename, engine="openpyxl", sheet_name="Tables", usecols="E:I"
        )
        .dropna()
        .astype({"ProductID": int, "CategoryID": int}),
        "dCategory": pd.read_excel(
            filename, engine="openpyxl", sheet_name="Tables", usecols="K:L"
        )
        .rename(columns={"CategoryID.1": "CategoryID"})
        .dropna()
        .astype({"CategoryID": int}),
    }
    for key, value in expected.items():
        assert_frame_equal(value, actual[key], check_dtype=False)
